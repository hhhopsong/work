# -*- coding: utf-8 -*-
"""
手动框选因子版：手动给定预测因子（90% 显著性框选）-> 逐步回归 -> 预测 + 出图

流程：
1. 在 MANUAL_FACTORS 中手动给定每个因子的 name / elem / month1 / month2 / lat / lon。
2. build_manual_predictors 按 (month1, month2) 分组，复用 predictor 计算各区域指数。
3. 把所有手动因子放入逐步回归，挑出最终入选因子。
4. 出图：
   - 区域图按要素分别出图，同一「时次（同要素 + 同月份组合）」的因子画在同一个子图。
   - 被逐步回归选中的因子用红色实线框；被剔除的因子用蓝色虚线框。
   - 叠加 90% 显著性打点（颜色 #757575）。
5. rollingCorr 面板 + 最终预报面板照旧。

注意：
- 该脚本依赖你原来的 climkit、cmaps、本地数据路径与字体环境。
- Cartopy 第一次使用 Natural Earth 数据可能联网下载海岸线。
"""

# =========================================================
# 0. imports
# =========================================================
from cartopy import crs as ccrs
import cartopy.feature as cfeature
from cartopy.mpl.ticker import LongitudeFormatter, LatitudeFormatter
from cartopy.util import add_cyclic_point

import matplotlib.pyplot as plt
from matplotlib import gridspec
from matplotlib.patches import Rectangle
from matplotlib.lines import Line2D
import cmaps

from scipy import stats
import xarray as xr
import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
import statsmodels.api as sm
from statsmodels.stats.outliers_influence import variance_inflation_factor

import os
import ast

from climkit.significance_test import corr_test, r_test
from climkit.data_read import *
from climkit.corr_reg import corr, regress
from climkit.lonlat_transform import transform


read_sic = sic


# =========================================================
# 1. 数据预处理
# =========================================================
def prepare_sic_dataset(ds):
    """Normalize HadISST sea-ice dataset to {'sic': [time, lat, lon]}."""
    rename_map = {}
    if 'latitude' in ds.coords and 'lat' not in ds.coords:
        rename_map['latitude'] = 'lat'
    if 'longitude' in ds.coords and 'lon' not in ds.coords:
        rename_map['longitude'] = 'lon'
    if len(rename_map) > 0:
        ds = ds.rename(rename_map)

    if 'sic' not in ds.data_vars:
        for cand in ('ice', 'siconc', 'ci'):
            if cand in ds.data_vars:
                ds = ds.rename({cand: 'sic'})
                break

    if 'sic' not in ds.data_vars:
        raise ValueError(f"SIC variable not found. Available vars: {list(ds.data_vars)}")

    sic_da = ds['sic']
    vmax = float(np.nanmax(sic_da.values))
    if vmax <= 1.5:
        sic_da = sic_da.where((sic_da >= 0) & (sic_da <= 1.0))
    else:
        sic_da = sic_da.where((sic_da >= 0) & (sic_da <= 100.0))

    out = xr.Dataset({'sic': sic_da})
    out = out.sortby('lat').sortby('lon')
    return out


def prepare_swvl_dataset(swvl_like):
    """Normalize soil moisture to {'swvl': [time, lat, lon]}. 支持 DataArray / Dataset"""
    if isinstance(swvl_like, xr.DataArray):
        da = swvl_like.copy()
        if da.name != 'swvl':
            da = da.rename('swvl')
        ds = da.to_dataset()
    elif isinstance(swvl_like, xr.Dataset):
        ds = swvl_like.copy()
        if 'swvl' not in ds.data_vars:
            for cand in ('swvl1', 'soil_moisture', 'sm'):
                if cand in ds.data_vars:
                    ds = ds.rename({cand: 'swvl'})
                    break
    else:
        raise TypeError("swvl_like must be xarray.DataArray or xarray.Dataset")

    rename_map = {}
    if 'latitude' in ds.coords and 'lat' not in ds.coords:
        rename_map['latitude'] = 'lat'
    if 'longitude' in ds.coords and 'lon' not in ds.coords:
        rename_map['longitude'] = 'lon'
    if len(rename_map) > 0:
        ds = ds.rename(rename_map)

    if 'swvl' not in ds.data_vars:
        raise ValueError(f"SWVL variable not found. Available vars: {list(ds.data_vars)}")

    out = xr.Dataset({'swvl': ds['swvl']})
    out = out.sortby('lat').sortby('lon')
    return out


# =========================================================
# 2. 辅助函数
# =========================================================
def _make_slice(coord, v1, v2):
    """根据坐标升降序自动生成 slice"""
    coord_values = coord.values
    if coord_values[0] < coord_values[-1]:
        return slice(min(v1, v2), max(v1, v2))
    else:
        return slice(max(v1, v2), min(v1, v2))


def month_to_season_order(month, cross_month=9):
    """把自然月映射到以 cross_month 为起点的季节年顺序。"""
    return ((month - cross_month) % 12) + 1


def select_region_no_cut(da, lat1, lat2, lon1, lon2):
    """
    支持普通经度区间和跨经线区间的区域选取。
    - 若 lon1 <= lon2: 普通区间
    - 若 lon1 >  lon2: 跨越经线，拼接两段
    """
    lat_sl = _make_slice(da.lat, lat1, lat2)

    if lon1 <= lon2:
        lon_sl = _make_slice(da.lon, lon1, lon2)
        return da.sel(lat=lat_sl, lon=lon_sl)

    lon_vals = da.lon.values
    part1 = da.sel(lat=lat_sl, lon=_make_slice(da.lon, lon1, np.nanmax(lon_vals)))
    part2 = da.sel(lat=lat_sl, lon=_make_slice(da.lon, np.nanmin(lon_vals), lon2))
    return xr.concat([part1, part2], dim='lon')


def rolling_corr_centered_index(x, y, window=11):
    """滑动相关，把索引从窗口右端年改成窗口中心年。"""
    rc = x.rolling(window=window).corr(y)
    shift_year = (window - 1) // 2
    if isinstance(rc.index, pd.DatetimeIndex):
        rc.index = rc.index - pd.DateOffset(years=shift_year)
    else:
        rc.index = rc.index - shift_year
    return rc


def get_corr_map_from_predictor_output(pred_out):
    """predictor 返回值转成 corr 字典。"""
    (
        _, _, _, TS, TS_pre, TS_all,
        t2mReg, t2mCorr,
        slpReg, slpCorr,
        sstReg, sstCorr,
        sicReg, sicCorr,
        swvlReg, swvlCorr
    ) = pred_out

    corr_map = {
        'sst': sstCorr,
        'slp': slpCorr,
        't2m': t2mCorr,
        'sic': sicCorr,
        'swvl': swvlCorr
    }
    return corr_map, TS, TS_pre, TS_all


# ---- 时期统计（用于 rollingCorr 面板上 orange 高亮）----
def calc_period_corr_stats(ts, x):
    sub = pd.concat([ts.rename('TS'), x.rename('X')], axis=1) \
        .replace([np.inf, -np.inf], np.nan).dropna()
    n = len(sub)
    if n < 3:
        return np.nan, np.nan, n
    try:
        model = smf.ols("TS ~ X", data=sub).fit()
        corr_val = sub['TS'].corr(sub['X'])
        pval = model.pvalues.get('X', np.nan)
    except Exception:
        corr_val, pval = np.nan, np.nan
    return corr_val, pval, n


def calc_late_rolling_stats(rolling_series, PR_time, r_sig, train_sign=None, min_count=3):
    s = rolling_series.copy()
    if isinstance(s.index, pd.DatetimeIndex):
        s_late = s[(s.index.year >= PR_time[0]) & (s.index.year <= PR_time[1])]
    else:
        s_late = s[(s.index >= PR_time[0]) & (s.index <= PR_time[1])]

    s_late = s_late.replace([np.inf, -np.inf], np.nan).dropna()
    late_n = len(s_late)
    if late_n < min_count:
        return {'late_n': late_n, 'late_sig_ratio': np.nan, 'late_same_sign_ratio': np.nan,
                'late_mean_corr': np.nan, 'late_sig_mean_corr': np.nan}

    s_sig = s_late[np.abs(s_late) > r_sig]
    late_sig_ratio = len(s_sig) / late_n if late_n > 0 else np.nan
    late_mean_corr = s_late.mean() if late_n > 0 else np.nan
    late_sig_mean_corr = s_sig.mean() if len(s_sig) > 0 else np.nan

    if (train_sign is None) or (len(s_sig) == 0):
        late_same_sign_ratio = np.nan
    else:
        late_same_sign_ratio = (np.sign(s_sig.values) == train_sign).mean()

    return {'late_n': late_n, 'late_sig_ratio': late_sig_ratio,
            'late_same_sign_ratio': late_same_sign_ratio,
            'late_mean_corr': late_mean_corr, 'late_sig_mean_corr': late_sig_mean_corr}


def enrich_meta_df_with_period_stats(meta_df, X_train_all_dict, X_pre_all_dict, X_roll_all_dict,
                                     TS, TS_pre, PR_time, rolling_window=11, alpha=0.1,
                                      p_value=0.1):
    """给 meta_df 添加训练/预测期及后期 rolling 显著性统计。"""
    meta_df = meta_df.copy()
    r_sig = r_test(rolling_window, alpha)

    cols = {k: [] for k in [
        'train_corr', 'train_pval', 'train_n', 'pre_corr', 'pre_pval', 'pre_n',
        'late_n', 'late_sig_ratio', 'late_same_sign_ratio', 'late_mean_corr', 'late_sig_mean_corr']}

    for _, row in meta_df.iterrows():
        xname = row['X_name']
        x_train = X_train_all_dict.get(xname, None)
        x_pre = X_pre_all_dict.get(xname, None)
        x_roll = X_roll_all_dict.get(xname, None)

        tr_corr, tr_pval, tr_n = calc_period_corr_stats(TS, x_train)
        cols['train_corr'].append(tr_corr); cols['train_pval'].append(tr_pval); cols['train_n'].append(tr_n)

        pr_corr, pr_pval, pr_n = calc_period_corr_stats(TS_pre, x_pre)
        cols['pre_corr'].append(pr_corr); cols['pre_pval'].append(pr_pval); cols['pre_n'].append(pr_n)

        train_sign = np.sign(tr_corr) if np.isfinite(tr_corr) and tr_corr != 0 else None
        late_stats = calc_late_rolling_stats(x_roll, PR_time, r_sig, train_sign=train_sign, min_count=3)
        for k in ['late_n', 'late_sig_ratio', 'late_same_sign_ratio', 'late_mean_corr', 'late_sig_mean_corr']:
            cols[k].append(late_stats[k])

    for k, v in cols.items():
        meta_df[k] = v

    meta_df['late_sig_same_sign'] = (
        (meta_df['pre_corr'] >= 0) &
        (meta_df['pre_pval'] <= p_value)
    )
    return meta_df


# =========================================================
# 3. predictor 主函数（保留原逻辑）
# =========================================================
def predictor(timeSerie, TR_time, PR_time, month1, month2=None, predictor_zone=None, cross_month=9):

    if predictor_zone is None:
        predictor_zone = []

    def seasonal_mean_same_year(da, months, start_year, end_year):
        return (da.sel(time=da["time.month"].isin(months))
                  .sel(time=slice(f"{start_year}", f"{end_year}"))
                  .groupby("time.year").mean("time")
                  .transpose("year", "lat", "lon"))

    def seasonal_mean_cross_year(da, months, start_year, end_year, cross_month=cross_month):
        da_sel = da.sel(time=da["time.month"].isin(months)).sel(
            time=slice(f"{start_year-1}", f"{end_year}"))
        season_year = xr.where(da_sel["time.month"] >= cross_month,
                               da_sel["time.year"] + 1, da_sel["time.year"])
        da_sel = da_sel.assign_coords(season_year=("time", season_year.data))
        out = (da_sel.groupby("season_year").mean("time")
               .rename({"season_year": "year"}).transpose("year", "lat", "lon"))
        return out.sel(year=slice(start_year, end_year))

    def get_seasonal_mean(da, months, start_year, end_year, cross_month=cross_month):
        cross_year = any(m >= cross_month for m in months)
        if cross_year:
            return seasonal_mean_cross_year(da, months, start_year, end_year, cross_month)
        else:
            return seasonal_mean_same_year(da, months, start_year, end_year)

    if month2 is None:
        t2m_imonth = get_seasonal_mean(t2m, month1, TR_time[0], TR_time[1], cross_month)
        slp_imonth = get_seasonal_mean(slp, month1, TR_time[0], TR_time[1], cross_month)
        sst_imonth = get_seasonal_mean(sst, month1, TR_time[0], TR_time[1], cross_month)
        sic_imonth = get_seasonal_mean(sic_ds, month1, TR_time[0], TR_time[1], cross_month)
        swvl_imonth = get_seasonal_mean(swvl, month1, TR_time[0], TR_time[1], cross_month)

        t2m_imonth_pre = get_seasonal_mean(t2m, month1, PR_time[0], PR_time[1], cross_month)
        slp_imonth_pre = get_seasonal_mean(slp, month1, PR_time[0], PR_time[1], cross_month)
        sst_imonth_pre = get_seasonal_mean(sst, month1, PR_time[0], PR_time[1], cross_month)
        sic_imonth_pre = get_seasonal_mean(sic_ds, month1, PR_time[0], PR_time[1], cross_month)
        swvl_imonth_pre = get_seasonal_mean(swvl, month1, PR_time[0], PR_time[1], cross_month)

        t2m_imonth_all = get_seasonal_mean(t2m, month1, TR_time[0], PR_time[1], cross_month)
        slp_imonth_all = get_seasonal_mean(slp, month1, TR_time[0], PR_time[1], cross_month)
        sst_imonth_all = get_seasonal_mean(sst, month1, TR_time[0], PR_time[1], cross_month)
        sic_imonth_all = get_seasonal_mean(sic_ds, month1, TR_time[0], PR_time[1], cross_month)
        swvl_imonth_all = get_seasonal_mean(swvl, month1, TR_time[0], PR_time[1], cross_month)

    else:
        def _mean(da, months):
            return (get_seasonal_mean(da, months, TR_time[0], TR_time[1], cross_month),
                    get_seasonal_mean(da, months, PR_time[0], PR_time[1], cross_month),
                    get_seasonal_mean(da, months, TR_time[0], PR_time[1], cross_month))

        t2m1, t2m1p, t2m1a = _mean(t2m, month1)
        slp1, slp1p, slp1a = _mean(slp, month1)
        sst1, sst1p, sst1a = _mean(sst, month1)
        sic1, sic1p, sic1a = _mean(sic_ds, month1)
        sw1, sw1p, sw1a = _mean(swvl, month1)

        t2m2, t2m2p, t2m2a = _mean(t2m, month2)
        slp2, slp2p, slp2a = _mean(slp, month2)
        sst2, sst2p, sst2a = _mean(sst, month2)
        sic2, sic2p, sic2a = _mean(sic_ds, month2)
        sw2, sw2p, sw2a = _mean(swvl, month2)

        t2m_imonth = t2m1 - t2m2
        slp_imonth = slp1 - slp2
        sst_imonth = sst1 - sst2
        sic_imonth = sic1 - sic2
        swvl_imonth = sw1 - sw2

        t2m_imonth_pre = t2m1p - t2m2p
        slp_imonth_pre = slp1p - slp2p
        sst_imonth_pre = sst1p - sst2p
        sic_imonth_pre = sic1p - sic2p
        swvl_imonth_pre = sw1p - sw2p

        t2m_imonth_all = t2m1a - t2m2a
        slp_imonth_all = slp1a - slp2a
        sst_imonth_all = sst1a - sst2a
        sic_imonth_all = sic1a - sic2a
        swvl_imonth_all = sw1a - sw2a

    timeSerie_train = timeSerie.sel(year=slice(f'{TR_time[0]}', f'{TR_time[1]}')).data
    train_years = pd.to_datetime(np.arange(TR_time[0], TR_time[1] + 1), format='%Y')
    pre_years = pd.to_datetime(np.arange(PR_time[0], PR_time[1] + 1), format='%Y')

    t2mReg, t2mCorr = regress(timeSerie_train, t2m_imonth['t2m'].data), corr(timeSerie_train, t2m_imonth['t2m'].data)
    slpReg, slpCorr = regress(timeSerie_train, slp_imonth['msl'].data), corr(timeSerie_train, slp_imonth['msl'].data)
    sstReg, sstCorr = regress(timeSerie_train, sst_imonth['sst'].data), corr(timeSerie_train, sst_imonth['sst'].data)
    sicReg, sicCorr = regress(timeSerie_train, sic_imonth['sic'].data), corr(timeSerie_train, sic_imonth['sic'].data)
    swvlReg, swvlCorr = regress(timeSerie_train, swvl_imonth['swvl'].data), corr(timeSerie_train, swvl_imonth['swvl'].data)

    t2mReg = xr.DataArray(t2mReg, coords=[t2m_imonth['lat'], t2m_imonth['lon']], dims=['lat', 'lon'], name='t2m_reg')
    slpReg = xr.DataArray(slpReg, coords=[slp_imonth['lat'], slp_imonth['lon']], dims=['lat', 'lon'], name='slp_reg')
    sstReg = xr.DataArray(sstReg, coords=[sst_imonth['lat'], sst_imonth['lon']], dims=['lat', 'lon'], name='sst_reg')
    sicReg = xr.DataArray(sicReg, coords=[sic_imonth['lat'], sic_imonth['lon']], dims=['lat', 'lon'], name='sic_reg')
    swvlReg = xr.DataArray(swvlReg, coords=[swvl_imonth['lat'], swvl_imonth['lon']], dims=['lat', 'lon'], name='swvl_reg')

    t2mCorr = xr.DataArray(t2mCorr, coords=[t2m_imonth['lat'], t2m_imonth['lon']], dims=['lat', 'lon'], name='t2m_corr')
    slpCorr = xr.DataArray(slpCorr, coords=[slp_imonth['lat'], slp_imonth['lon']], dims=['lat', 'lon'], name='slp_corr')
    sstCorr = xr.DataArray(sstCorr, coords=[sst_imonth['lat'], sst_imonth['lon']], dims=['lat', 'lon'], name='sst_corr')
    sicCorr = xr.DataArray(sicCorr, coords=[sic_imonth['lat'], sic_imonth['lon']], dims=['lat', 'lon'], name='sic_corr')
    swvlCorr = xr.DataArray(swvlCorr, coords=[swvl_imonth['lat'], swvl_imonth['lon']], dims=['lat', 'lon'], name='swvl_corr')

    nor_mean = np.mean(timeSerie_train)
    nor_std = np.std(timeSerie_train)
    timeSerie_train = (timeSerie_train - nor_mean) / nor_std
    TS = pd.Series(timeSerie_train, index=train_years, name='TS')

    timeSerie_pre = timeSerie.sel(year=slice(f'{PR_time[0]}', f'{PR_time[1]}'))
    timeSerie_pre = (timeSerie_pre - nor_mean) / nor_std
    TS_pre = pd.Series(timeSerie_pre, index=pre_years, name='TS_pre')

    timeSerie_all = timeSerie.sel(year=slice(f'{TR_time[0]}', f'{PR_time[1]}'))
    timeSerie_all = (timeSerie_all - nor_mean) / nor_std
    TS_all = pd.Series(timeSerie_all, index=np.arange(TR_time[0], PR_time[1] + 1), name='TS_all')

    field_map = {
        'sst': {'corr': sstCorr, 'train': sst_imonth, 'pre': sst_imonth_pre, 'all': sst_imonth_all, 'var': 'sst'},
        'slp': {'corr': slpCorr, 'train': slp_imonth, 'pre': slp_imonth_pre, 'all': slp_imonth_all, 'var': 'msl'},
        't2m': {'corr': t2mCorr, 'train': t2m_imonth, 'pre': t2m_imonth_pre, 'all': t2m_imonth_all, 'var': 't2m'},
        'sic': {'corr': sicCorr, 'train': sic_imonth, 'pre': sic_imonth_pre, 'all': sic_imonth_all, 'var': 'sic'},
        'swvl': {'corr': swvlCorr, 'train': swvl_imonth, 'pre': swvl_imonth_pre, 'all': swvl_imonth_all, 'var': 'swvl'}
    }

    X_train_dict = {}
    X_pre_dict = {}
    X_rollingCorr_dict = {}

    index = 0
    for izone in predictor_zone:
        index += 1
        elem = izone[0].lower()
        X_zone = izone[1:]

        if elem not in field_map:
            raise ValueError(f"Unsupported predictor element: {elem}.")

        corr_da_ = field_map[elem]['corr']
        train_da_ = field_map[elem]['train']
        pre_da_ = field_map[elem]['pre']
        all_da_ = field_map[elem]['all']
        var_name = field_map[elem]['var']

        lat1, lat2, lon1_raw, lon2_raw = X_zone

        if np.nanmax(corr_da_.lon.values) > 180:
            lon1 = lon1_raw % 360
            lon2 = lon2_raw % 360
        else:
            lon1 = lon1_raw
            lon2 = lon2_raw
            if lon1 > 180:
                lon1 = ((lon1 + 180) % 360) - 180
            if lon2 > 180:
                lon2 = ((lon2 + 180) % 360) - 180

        weight = select_region_no_cut(corr_da_, lat1, lat2, lon1, lon2)
        sig_mask = np.abs(weight) > r_test(TR_time[1] - TR_time[0] + 1, 0.1)

        X_train = select_region_no_cut(train_da_[var_name], lat1, lat2, lon1, lon2) * xr.where(sig_mask, weight, np.nan)
        X_train = X_train.mean(['lat', 'lon'])
        X_mean, X_std = X_train.mean(), X_train.std()
        X_train = (X_train - X_mean) / X_std
        X_train = pd.Series(X_train.data, index=train_years, name=f'X{index}_train')

        X_pre = select_region_no_cut(pre_da_[var_name], lat1, lat2, lon1, lon2) * xr.where(sig_mask, weight, np.nan)
        X_pre = X_pre.mean(['lat', 'lon'])
        X_pre = (X_pre - X_mean) / X_std
        X_pre = pd.Series(X_pre.data, index=pre_years, name=f'X{index}_pre')

        X_all = select_region_no_cut(all_da_[var_name], lat1, lat2, lon1, lon2) * xr.where(sig_mask, weight, np.nan)
        X_all = X_all.mean(['lat', 'lon'])
        X_all = (X_all - X_mean) / X_std
        X_all = pd.Series(X_all.data, index=np.arange(TR_time[0], PR_time[1] + 1), name=f'X{index}_all')

        X_rollingCorr = rolling_corr_centered_index(X_all, TS_all, window=11)
        X_rollingCorr.name = f'X{index}_rollingCorr'

        X_train_dict[f'X{index}'] = X_train
        X_pre_dict[f'X{index}'] = X_pre
        X_rollingCorr_dict[f'X{index}'] = X_rollingCorr

    return (X_train_dict, X_pre_dict, X_rollingCorr_dict,
            TS, TS_pre, TS_all,
            t2mReg, t2mCorr, slpReg, slpCorr, sstReg, sstCorr,
            sicReg, sicCorr, swvlReg, swvlCorr)


# =========================================================
# 4. 手动因子定义与构造
# =========================================================
# 字段：name, elem, month1(list), month2(list 或 None), lat1, lat2, lon1, lon2
# 经度统一用 [-180, 180]：E 为正，W 为负。lon1 > lon2 表示跨经线。
MANUAL_FACTORS = [
    ('X1_1',  't2m',  [3, 4], [11, 12],  45, 70,   20,  140),
    ('X1_2',  't2m',  [4, 5], [11, 12],  45, 70,   20,  140),
    ('X1_3',  't2m',  [5, 6], [11, 12],  45, 70,   70,  140),
    ('X1_4',  't2m',  [5, 6], [3, 4],    55, 75,   95,  130),
    ('X1_5',  't2m',  [3, 4], None,      40, 60,   20,   55),
    ('X1_6',  't2m',  [5, 6], None,      45, 73,   75,   130),
    # # La Nina trend
    ('X2_1',  'sst',  [3, 4], [11, 12], -13,  9,  170,  -80),
    ('X2_2',  'sst',  [3, 4], [12, 1],  -13,  9,  170,  -80),
    ('X2_3',  'sst',  [3, 4], [1, 2],   -13,  9,  170,  -80),
    ('X2_4', 'sst',   [4, 5], [11, 12], -13,  9,  170,  -80),
    ('X2_5', 'sst',   [4, 5], [12, 1],  -13,  9,  170,  -80),
    ('X2_6', 'sst',   [4, 5], [1, 2],   -13,  9,  170,  -80),
    ('X2_7', 'sst',   [4, 5], [2, 3],   -13,  9,  170,  -80),
    ('X2_8', 'sst',   [5, 6], [11, 12], -13,  9,  125,  -80),
    ('X2_9', 'sst',   [5, 6], [12, 1],  -13,  9,  125,  -80),
    ('X2_10', 'sst',  [5,  6], [1, 2],   -13,  9,  125,  -80),
    ('X2_11', 'sst',  [5, 6], [2, 3],   -13,  9,  170,  -80),
    ('X2_12', 'sst',  [5, 6], [3, 4],   -13,  9,  125,  -80),
    # # La Nina mean
    ('X2_13', 'sst',  [3, 4], None,     -10, 10, -160,  -80),
    ('X2_14', 'sst',  [4, 5], None,     -10, 10, -160,  -80),
    ('X2_15', 'sst',  [5, 6], None,     -10, 10, -160,  -80),
    # # NT
    # ## single
    ('X2_16', 'sst',  [4, 5], [12, 1],   30, 50,  -45,  -10),
    ('X2_17', 'sst',  [5, 6], [11, 12],  30, 50,  -45,  -10),
    ('X2_18', 'sst',  [5, 6], [12, 1],   30, 50,  -45,  -10),
    ('X2_19', 'sst',  [5, 6], [1, 2],    30, 50,  -45,  -10),
    # # # ## tripole
    ('X2_20', 'sst',  [4, 5], [12, 1],   18, 70,  -80,  -10),
    ('X2_21', 'sst',  [4, 5], [1, 2],    18, 70,  -80,  -10),
    ('X2_22', 'sst',  [5, 6], [12, 1],   18, 70,  -80,    0),
    ('X2_23', 'sst',  [5, 6], [1, 2],    15, 70,  -80,  -10),
    ('X2_24', 'sst',  [5, 6], [2, 3],    15, 70,  -80,  -10),
    ('X2_25', 'sst',  [5, 6], [3, 4],    5,  70,  -80,  -10),
    # # # La Nina single trend
    ('X2_26', 'sst', [4], [12], -13, 9, 170, -80),
    ('X2_27', 'sst', [4], [2], -13, 9, 170, -80),
    ('X2_28', 'sst', [4], [3], -13, 9, 170, -80),
    ('X2_29', 'sst', [5], [12], -13, 9, 170, -80),
    ('X2_30', 'sst', [5], [1], -13, 9, 170, -80),
    ('X2_31', 'sst', [5], [2], -13, 9, 170, -80),
    ('X2_32', 'sst', [5], [3], -13, 9, 170, -80),
    ('X2_33', 'sst', [5], [4], -13, 9, 170, -80),
    ('X2_34', 'sst', [6], [1], -13, 9, 140, -80),
    ('X2_35', 'sst', [6], [2], -13, 9, 170, -80),
    ('X2_36', 'sst', [6], [3], -13, 9, 140, -80),
    ('X2_37', 'sst', [6], [4], -13, 9, 170, -80),
    # # # NT single trend
    ('X2_38', 'sst', [5], [4],  18,38, -90, -40),
    ('X2_39', 'sst', [6], [3],  28,46, -90, -20),
    ('X2_40', 'sst', [6], [4],  28,46, -90, -20),
    ('X2_41', 'sst', [6], [5],  28,50, -90, -20),
    # # WP
    ('X2_42', 'sst',  [3, 4], [11, 12], -30,  15,  140,  180),
    ('X2_43', 'sst',  [4, 5], [11, 12], -30,  15,  140,  180),
    ('X2_44', 'sst',  [4, 5], [12,  1], -30,  15,  140,  180),
    ('X2_45', 'sst',  [5, 6], [11, 12], -30,  15,  140,  180),
    ('X2_46', 'sst',  [5, 6], [12,  1], -30,  15,  140,  180),
    ('X2_47', 'sst',  [4], [11], -30,  20,  140,  180),
    ('X2_48', 'sst',  [4], [12], -30,  20,  140,  180),
    ('X2_49', 'sst',  [5], [11], -30,  20,  140,  180),
    ('X2_50', 'sst',  [5], [12], -30,  20,  140,  180),
    ('X2_51', 'sst',  [6], [11], -30,  30,  140,  180),
    ('X2_52', 'sst',  [6], [12], -30,  30,  140,  180),
    ('X2_53', 'sst',  [6],  [ 1], -30,  20,  140,  180),
    ('X2_54', 'sst',  [6], [ 3], -10,  10,  140,  180),
    # # EP Ln Nina
    ('X2_55', 'sst',  [4, 5], [2, 3], -10,  10, -125,  -80),
    ('X2_56', 'sst',  [5, 6], [12,1], -10,  10, -125,  -80),
    ('X2_57', 'sst',  [5, 6], [1, 2], -10,  10, -125,  -80),
    ('X2_58', 'sst',  [5, 6], [2, 3], -10,  10, -125,  -80),
    ('X2_59', 'sst',  [5, 6], [2, 3], -10,  10, -125,  -80),
    ('X2_60', 'sst',  [ 5], [ 2], -10,  10, -125,  -80),
    ('X2_61', 'sst',  [ 5], [ 3], -10,  10, -125,  -80),
    ('X2_62', 'sst',  [ 5], [ 4], -10,  10, -125,  -80),
    ('X2_63', 'sst',  [ 6], [ 1], -10,  10, -125,  -80),
    ('X2_64', 'sst',  [ 6], [ 2], -10,  10, -125,  -80),
    # # NAO
    ('X3_1', 'slp',  [4, 5], [12, 1],   18, 85,  -95,    0),
    ('X3_2', 'slp',  [5, 6], [11, 12],  20, 85, -140,    0),
    ('X3_3', 'slp',  [5, 6], [12, 1],   30, 85, -140,  -10),
    ('X3_4', 'slp',  [5, 6], [1, 2],    40, 60,  -50,   40),
    ('X3_5', 'slp',  [5, 6], [2, 3],    35, 70,  -50,   60),
    ('X3_6', 'slp',  [5, 6], None,      22, 85,  -80,    0),
    # # NAO single trend
    ('X3_7', 'slp',  [4], [2],  35,  70,  -40,  0),
    ('X3_8', 'slp',  [5], [2],  30,  68,  -40,  40),
    ('X3_9', 'slp',  [6], [12], 22,  60,  -40,  40), # NT SLP
    ('X3_10','slp',  [6], [2], 22,  60,  -50,  10),
    #
    ('X4_1', 'sic',  [4, 5], [12, 1],   70, 90,   90,  -90),
    ('X5_1', 'swvl', [3, 4], None,      52, 68,   20,   75),
    # # W SM trend
    ('X5_2', 'swvl', [3, 4], [11, 12],  50,  68,  20,  90),
    ('X5_3', 'swvl', [3, 4], [12,  1],  50,  68,  20,  90),
    ('X5_4', 'swvl', [3, 4], [ 1,  2],  50,  68,  20,  90),
    ('X5_5', 'swvl', [ 3], [12],  50,  68,  20,  90),
    ('X5_6', 'swvl', [ 3], [ 1],  50,  68,  20,  90),
    ('X5_7', 'swvl', [ 3], [ 2],  50,  68,  20,  90),
    # # E SM trend
    ('X5_8', 'swvl', [5, 6], [11, 12],  50,  68,  90,  136),
    ('X5_9', 'swvl', [5, 6], [12,  1],  50,  68,  90,  136),
    ('X5_10', 'swvl',[5, 6], [ 1,  2],  50,  68,  90,  136),
    ('X5_11', 'swvl',[5, 6], [ 2,  3],  50,  68,  90,  136),
    ('X5_12', 'swvl',[5, 6], [ 3,  4],  50,  68,  90,  136),
    ('X5_13', 'swvl',[6], [1],  50,  68,  90,  136),
    ('X5_14', 'swvl',[6], [2],  50,  68,  90,  136),
    ('X5_15', 'swvl',[6], [3],  50,  68,  90,  136),
    ('X5_16', 'swvl',[6], [4],  50,  68,  90,  136),
    ('X5_17', 'swvl',[6], [5],  50,  68,  90,  136),
    # # E SM mean
    ('X5_18', 'swvl',[5, 6], None,  50,  68,  96,  136),
]


def make_month_expr(month1, month2):
    """生成与你框选一致的月份表达式，如 3+4 或 (5+6)-(11+12)。"""
    def fmt(ms):
        return '+'.join(str(m) for m in ms)
    if month2 is None:
        return fmt(month1)
    return f'({fmt(month1)})-({fmt(month2)})'


def build_manual_predictors(timeSerie, TR_time, PR_time, manual_factors, cross_month=9):
    """
    按 (month1, month2) 分组，复用 predictor 计算每个手动框选区域的指数。
    返回 X_train / X_pre / X_roll 字典（key 为你给定的 X 名称）以及 meta_df。
    """
    # 按月份组合分组（同一时次的所有要素一起算）
    groups = {}
    for f in manual_factors:
        name, elem, m1, m2, lat1, lat2, lon1, lon2 = f
        key = (tuple(m1), tuple(m2) if m2 is not None else None)
        groups.setdefault(key, []).append(f)

    X_train_all, X_pre_all, X_roll_all = {}, {}, {}
    meta_records = []
    TS_ref = TS_pre_ref = TS_all_ref = None

    for (m1, m2), flist in groups.items():
        month1 = list(m1)
        month2 = list(m2) if m2 is not None else None

        zones = [[f[1], f[4], f[5], f[6], f[7]] for f in flist]

        pred_out = predictor(
            timeSerie=timeSerie, TR_time=TR_time, PR_time=PR_time,
            month1=month1, month2=month2, predictor_zone=zones, cross_month=cross_month
        )
        X_train_dict, X_pre_dict, X_roll_dict, TS, TS_pre, TS_all = pred_out[:6]

        if TS_ref is None:
            TS_ref, TS_pre_ref, TS_all_ref = TS, TS_pre, TS_all

        for i, f in enumerate(flist, start=1):
            name = f[0]
            old_key = f'X{i}'
            X_train_all[name] = X_train_dict[old_key].rename(name)
            X_pre_all[name] = X_pre_dict[old_key].rename(name)
            X_roll_all[name] = X_roll_dict[old_key].rename(name)

            meta_records.append({
                'X_name': name,
                'elem': f[1],
                'month1': tuple(month1),
                'month2': tuple(month2) if month2 is not None else None,
                'month_expr': make_month_expr(month1, month2),
                'lat1': f[4], 'lat2': f[5], 'lon1': f[6], 'lon2': f[7],
            })

    meta_df = pd.DataFrame(meta_records)
    return X_train_all, X_pre_all, X_roll_all, meta_df, TS_ref, TS_pre_ref, TS_all_ref


# =========================================================
# 5. 逐步回归工具
# =========================================================
def calc_vif(df, features):
    if len(features) <= 1:
        return pd.Series([1.0] * len(features), index=features, dtype=float)

    X = df[features].copy().replace([np.inf, -np.inf], np.nan).dropna()
    if len(X) <= len(features):
        return pd.Series([1.0] * len(features), index=features, dtype=float)

    X_const = sm.add_constant(X, has_constant='add')
    vif_values = []
    for i in range(1, X_const.shape[1]):
        try:
            vif_values.append(variance_inflation_factor(X_const.values, i))
        except Exception:
            vif_values.append(np.nan)

    out = pd.Series(vif_values, index=features, dtype=float)
    out = out.replace([np.inf, -np.inf], np.nan).fillna(999.0)
    return out


def stepwise_regression(data, Dep, InDep, limit):
    """
    后退逐步回归：
    从全部候选因子开始建模，每次剔除 p 值最大的因子，
    直到所有因子的 p 值 <= limit。

    Parameters
    ----------
    data : pd.DataFrame
        包含因变量和自变量的数据。
    Dep : str
        因变量列名。
    InDep : list
        自变量列名列表。
    limit : float
        p 值剔除阈值。

    Returns
    -------
    result : statsmodels RegressionResults or None
        最终 OLS 回归结果。
    selected : list
        最终保留的自变量名称。
    """
    InDep = InDep.copy()

    for i in range(len(InDep)):
        if len(InDep) == 0:
            return None, []

        data_model = data[[Dep] + InDep].replace([np.inf, -np.inf], np.nan).dropna()

        x = sm.add_constant(data_model[InDep], has_constant='add')
        y = data_model[Dep]

        model = sm.OLS(y, x)
        result = model.fit()

        pvalues = result.pvalues.copy()
        pvalues.drop('const', inplace=True, errors='ignore')
        pvalues.drop('Intercept', inplace=True, errors='ignore')

        if len(pvalues) == 0:
            return result, InDep

        pmax = max(pvalues)

        if pmax > limit:
            ind = pvalues.idxmax()
            InDep.remove(ind)
            print(f"[Backward-p] remove {ind}, p={pmax:.4f}")
        else:
            return result, InDep

    if len(InDep) == 0:
        return None, []

    data_model = data[[Dep] + InDep].replace([np.inf, -np.inf], np.nan).dropna()
    x = sm.add_constant(data_model[InDep], has_constant='add')
    y = data_model[Dep]
    result = sm.OLS(y, x).fit()

    return result, InDep


# =========================================================
# 6. 区域图：手动框选因子（红=入选，蓝虚线=被剔除）
# =========================================================
ELEM_LABEL = {
    'sst': 'SST', 'sic': 'SIC', 'swvl': 'Soil moisture',
    'slp': 'SLP', 't2m': '2-m temperature'
}


def set_sci_map_style():
    plt.rcParams.update({
        'font.family': ['Arial', 'Helvetica', 'DejaVu Sans'],
        'axes.unicode_minus': False,
        'font.size': 8, 'axes.labelsize': 8, 'axes.titlesize': 9,
        'xtick.labelsize': 7, 'ytick.labelsize': 7, 'legend.fontsize': 7,
        'figure.titlesize': 11, 'axes.linewidth': 0.6,
        'pdf.fonttype': 42, 'ps.fonttype': 42, 'savefig.dpi': 600
    })


def _safe_month_list(x):
    """支持 tuple/list/None/字符串形式的 month。"""
    if x is None:
        return None
    if isinstance(x, float) and np.isnan(x):
        return None
    if isinstance(x, str):
        xs = x.strip()
        if xs.lower() in ['none', 'nan', '']:
            return None
        try:
            x = ast.literal_eval(xs)
        except Exception:
            x = xs.replace('[', '').replace(']', '').replace('(', '').replace(')', '')
            return [int(v.strip()) for v in x.split(',') if v.strip() != '']
    if isinstance(x, (list, tuple, np.ndarray)):
        return [int(v) for v in x]
    return [int(x)]


def _month_sort_key(row, cross_month=9):
    """按气候季节顺序给月份组合排序。"""
    m1 = _safe_month_list(row['month1'])
    m2 = _safe_month_list(row['month2'])
    key = []
    if m1 is not None:
        key.extend([month_to_season_order(m, cross_month) for m in m1])
    if m2 is not None:
        key.extend([month_to_season_order(m, cross_month) for m in m2])
    return tuple(key)


def _lon_to_data(lon, lon_values):
    """把 lon 转换到当前数据经度体系。"""
    lon = float(lon)
    if np.nanmax(lon_values) > 180:
        return lon % 360
    else:
        return ((lon + 180) % 360) - 180


def draw_factor_box(ax, row, lon_values, selected=True, label=True, linewidth=1.0, zorder=20):
    """
    画因子区域框。
    - selected=True : 红色实线（逐步回归入选）
    - selected=False: 蓝色虚线（逐步回归剔除）
    """
    lat_low = min(float(row['lat1']), float(row['lat2']))
    lat_high = max(float(row['lat1']), float(row['lat2']))

    lon1 = _lon_to_data(row['lon1'], lon_values)
    lon2 = _lon_to_data(row['lon2'], lon_values)

    lon_min = float(np.nanmin(lon_values))
    lon_max = float(np.nanmax(lon_values))

    wraps = (lon1 > lon2)

    if selected:
        edgecolor = 'red'
        linestyle = '-'
        lw = linewidth + 0.4
    else:
        edgecolor = 'blue'
        linestyle = (0, (4, 2))
        lw = linewidth

    if not wraps:
        segments = [(lon1, lon2)]
    else:
        segments = [(lon1, lon_max), (lon_min, lon2)]

    for a, b in segments:
        if b < a:
            continue
        rect = Rectangle(
            xy=(a, lat_low), width=b - a, height=lat_high - lat_low,
            fill=False, edgecolor=edgecolor, linewidth=lw, linestyle=linestyle,
            transform=ccrs.PlateCarree(), zorder=zorder
        )
        ax.add_patch(rect)

    if label:
        ax.text(
            lon1, lat_high, str(row['X_name']),
            transform=ccrs.PlateCarree(), ha='left', va='bottom',
            fontsize=6, color=edgecolor, fontweight='bold' if selected else 'normal',
            bbox=dict(boxstyle='round,pad=0.12', fc='white', ec='none', alpha=0.65),
            zorder=zorder + 1
        )


def get_corr_da_for_month_expr(timeSerie, TR_time, PR_time, elem, month1, month2, cross_month=9):
    """针对某个 elem + 月份组合重新计算相关场。"""
    pred_out = predictor(
        timeSerie=timeSerie, TR_time=TR_time, PR_time=PR_time,
        month1=month1, month2=month2, predictor_zone=[], cross_month=cross_month
    )
    corr_map, _, _, _ = get_corr_map_from_predictor_output(pred_out)
    return corr_map[elem].sortby('lat').sortby('lon')


def add_significance_stippling(ax, corr_da, alpha=0.1, color='#757575', size=1.0,
                               stride=1, min_lat=None, alpha_scatter=0.65, zorder=12):
    """相关场上添加显著性打点（默认 90% 显著性）。"""
    corr_da = corr_da.sortby('lat').sortby('lon')
    r_crit = r_test(TR_time[1] - TR_time[0] + 1, alpha)

    arr = corr_da.values
    lats = corr_da['lat'].values
    lons = corr_da['lon'].values

    sig_mask = np.isfinite(arr) & (np.abs(arr) > r_crit)
    if min_lat is not None:
        sig_mask = sig_mask & (lats[:, None] >= float(min_lat))
    if stride is not None and stride > 1:
        ii, jj = np.indices(sig_mask.shape)
        sig_mask = sig_mask & (ii % stride == 0) & (jj % stride == 0)
    if not np.any(sig_mask):
        return

    lon2d, lat2d = np.meshgrid(lons, lats)
    ax.scatter(
        lon2d[sig_mask], lat2d[sig_mask], s=size, c=color, marker='.',
        linewidths=0, alpha=alpha_scatter, transform=ccrs.PlateCarree(),
        zorder=zorder, rasterized=True
    )


def plot_manual_predictor_region_maps(
    meta_df, timeSerie, TR_time, PR_time, out_dir, selected_predictors,
    elements=('sst', 'sic', 'swvl', 'slp', 't2m'),
    cross_month=9, ncols=3, figsize_per_panel=(2.6, 1.8),
    cmap='RdBu_r', levels=np.arange(-1.0, 1.01, 0.1),
    label_xname=True, dpi=600,
    stipple_sig=True, stipple_alpha=0.1, stipple_color='#757575',
    stipple_size=1.0, stipple_stride=1, stipple_min_lat=None,
    wspace=0.06, hspace=0.14
):
    """
    按要素分别出图；同一「时次（同要素 + 同月份组合）」的因子画在同一个子图。
    红实线框=入选，蓝虚线框=被剔除。
    """
    set_sci_map_style()
    os.makedirs(out_dir, exist_ok=True)

    selected_set = set(selected_predictors)
    saved_files = []

    if len(meta_df) == 0:
        print('[Region maps] meta_df is empty. Skip.')
        return saved_files

    for elem in elements:
        sub_all = meta_df[meta_df['elem'].astype(str).str.lower() == elem.lower()].copy()
        if len(sub_all) == 0:
            continue

        sub_all['_order'] = sub_all.apply(lambda r: _month_sort_key(r, cross_month), axis=1)
        sub_all['_grpkey'] = sub_all.apply(lambda r: (str(r['month1']), str(r['month2'])), axis=1)

        grp_order = (sub_all.sort_values('_order')
                     .drop_duplicates('_grpkey')['_grpkey'].tolist())

        n_panel = len(grp_order)
        ncols_use = min(ncols, n_panel)
        nrows_use = int(np.ceil(n_panel / ncols_use))

        fig_w = figsize_per_panel[0] * ncols_use
        fig_h = figsize_per_panel[1] * nrows_use + 0.7

        fig, axes = plt.subplots(
            nrows_use, ncols_use, figsize=(fig_w, fig_h),
            subplot_kw={'projection': ccrs.PlateCarree(central_longitude=180)},
            constrained_layout=False
        )
        axes = np.atleast_1d(axes).ravel()
        mappable = None

        for ipanel, gk in enumerate(grp_order):
            ax = axes[ipanel]
            sub = sub_all[sub_all['_grpkey'] == gk].copy()
            row0 = sub.iloc[0]

            month1 = _safe_month_list(row0['month1'])
            month2 = _safe_month_list(row0['month2'])

            corr_da = get_corr_da_for_month_expr(
                timeSerie, TR_time, PR_time, elem, month1, month2, cross_month
            ).sortby('lat').sortby('lon')

            corr_plot, lon_plot = add_cyclic_point(corr_da.values, coord=corr_da['lon'].values)
            lat_plot = corr_da['lat'].values

            mappable = ax.contourf(
                lon_plot, lat_plot, corr_plot, levels=levels, cmap=cmap,
                extend='both', transform=ccrs.PlateCarree(), zorder=1
            )

            if stipple_sig:
                add_significance_stippling(
                    ax=ax, corr_da=corr_da, alpha=stipple_alpha, color=stipple_color,
                    size=stipple_size, stride=stipple_stride, min_lat=stipple_min_lat, zorder=12
                )

            ax.coastlines(resolution='110m', linewidth=0.42, zorder=15)
            ax.set_global()
            ax.set_xticks(np.arange(-180, 181, 60), crs=ccrs.PlateCarree())
            ax.set_yticks(np.arange(-60, 91, 30), crs=ccrs.PlateCarree())
            ax.xaxis.set_major_formatter(LongitudeFormatter())
            ax.yaxis.set_major_formatter(LatitudeFormatter())

            row_id, col_id = ipanel // ncols_use, ipanel % ncols_use
            if row_id != nrows_use - 1:
                ax.set_xticklabels([])
            if col_id != 0:
                ax.set_yticklabels([])
            ax.tick_params(length=1.8, width=0.45, pad=1.5)

            for _, rg in sub.iterrows():
                draw_factor_box(
                    ax=ax, row=rg, lon_values=corr_da['lon'].values,
                    selected=(rg['X_name'] in selected_set),
                    label=label_xname, linewidth=1.0, zorder=20
                )

            panel_label = chr(ord('a') + ipanel)
            ax.set_title(f'({panel_label}) {row0["month_expr"]}  n={len(sub)}',
                         loc='left', fontsize=7.2, pad=2)

        for j in range(n_panel, len(axes)):
            fig.delaxes(axes[j])

        fig.subplots_adjust(left=0.04, right=0.99, top=0.91, bottom=0.11,
                            wspace=wspace, hspace=hspace)

        cbar = fig.colorbar(mappable, ax=axes[:n_panel], orientation='horizontal',
                            fraction=0.032, pad=0.045, aspect=45)
        cbar.set_label('Correlation coefficient', fontsize=8)
        cbar.ax.tick_params(labelsize=7, length=2.0, width=0.45)

        legend_elems = [
            Line2D([0], [0], color='red', lw=1.4, linestyle='-', label='Selected (stepwise)'),
            Line2D([0], [0], color='blue', lw=1.2, linestyle=(0, (4, 2)), label='Removed (stepwise)')
        ]
        fig.legend(handles=legend_elems, loc='upper right', fontsize=7,
                   frameon=False, ncol=2, bbox_to_anchor=(0.99, 0.995))

        fig.suptitle(f'{ELEM_LABEL.get(elem, elem.upper())} predictor regions',
                     y=0.985, x=0.4, fontsize=10.5)

        png_path = os.path.join(out_dir, f'manual_regions_{elem}.png')
        pdf_path = os.path.join(out_dir, f'manual_regions_{elem}.pdf')
        fig.savefig(png_path, dpi=dpi, bbox_inches='tight')
        fig.savefig(pdf_path, bbox_inches='tight')
        plt.close(fig)

        saved_files.extend([png_path, pdf_path])
        print(f'[Saved] {png_path}')
        print(f'[Saved] {pdf_path}')

    return saved_files


# =========================================================
# 7. 数据读取
# =========================================================
PYFILE = r"/volumes/TiPlus7100/PyFile"
DATA = r"/volumes/TiPlus7100/data"

EHCI = xr.open_dataset(f"{PYFILE}/p5/data/EHCI_daily.nc")
EHCI = EHCI.groupby('time.year')
EHCI30 = EHCI.apply(lambda x: (x > 0.5).sum())
EHCI30 = (EHCI30 - EHCI30.mean()) / EHCI30.std('year')
EHCI30 = EHCI30['EHCI']

# 2mT
t2m = era5_land(fr"{DATA}/ERA5/ERA5_land/uv_2mTTd_sfp_pre_0.nc", 1961, 2022, 't2m')

# SLP
slp = era5_s(fr"{DATA}/ERA5/ERA5_singleLev/ERA5_sgLEv.nc", 1961, 2022, 'msl')

# SST
sst = ersst(fr"{DATA}/NOAA/ERSSTv5/sst.mnmean.nc", 1961, 2022)

# SIC
sic_ds = prepare_sic_dataset(read_sic(fr"{DATA}/NOAA/HadISST/HadISST_ice.nc", 1961, 2022))

# SWVL = swvl1 + swvl2
swvl1 = era5_land(fr"{DATA}/ERA5/ERA5_land/sm.nc", 1961, 2022, 'swvl1')
swvl2 = era5_land(fr"{DATA}/ERA5/ERA5_land/sm.nc", 1961, 2022, 'swvl2')
swvl1_da = swvl1['swvl1'] if isinstance(swvl1, xr.Dataset) else swvl1
swvl2_da = swvl2['swvl2'] if isinstance(swvl2, xr.Dataset) else swvl2
swvl = prepare_swvl_dataset((swvl1_da + swvl2_da).rename('swvl'))

TR_time = [1962, 2006]
PR_time = [2007, 2022]
timeSerie = EHCI30
cross_month = 9

slope, intercept_trend, r_value, p_value, std_err = stats.linregress(
    [i for i in range(len(EHCI30))], EHCI30
)
print(f"##################################{p_value}#########################################")


# =========================================================
# 8. 构造手动因子
# =========================================================
X_train_all_dict, X_pre_all_dict, X_roll_all_dict, meta_df, TS, TS_pre, TS_all = build_manual_predictors(
    timeSerie=timeSerie, TR_time=TR_time, PR_time=PR_time,
    manual_factors=MANUAL_FACTORS, cross_month=cross_month
)

# 给 meta_df 补充时期统计（供 rollingCorr 面板 orange 高亮）
meta_df = enrich_meta_df_with_period_stats(
    meta_df=meta_df,
    X_train_all_dict=X_train_all_dict,
    X_pre_all_dict=X_pre_all_dict,
    X_roll_all_dict=X_roll_all_dict,
    TS=TS, TS_pre=TS_pre, PR_time=PR_time,
    rolling_window=11, alpha=0.1, p_value=0.1
)

print(f'共有 {len(X_train_all_dict)} 个手动框选预测因子')
print(meta_df)

meta_df.to_csv(fr"{PYFILE}/p5/data/manual_predictor_catalog.csv",
               index=False, encoding='utf-8-sig')


# =========================================================
# 9. 因子间及 EHCI 相关系数表 + 导出 Excel
# =========================================================
from scipy import stats
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def calc_corr_p_n(df):
    """
    计算 Pearson 相关系数、p 值、样本数矩阵。
    df: columns = ['EHCI', X1, X2, ...]
    """
    cols = df.columns.tolist()

    corr_mat = pd.DataFrame(np.nan, index=cols, columns=cols)
    p_mat = pd.DataFrame(np.nan, index=cols, columns=cols)
    n_mat = pd.DataFrame(0, index=cols, columns=cols)

    for i, c1 in enumerate(cols):
        for j, c2 in enumerate(cols):
            sub = df[[c1, c2]].replace([np.inf, -np.inf], np.nan).dropna()
            n = len(sub)
            n_mat.loc[c1, c2] = n

            if c1 == c2:
                corr_mat.loc[c1, c2] = 1.0
                p_mat.loc[c1, c2] = 0.0
            elif n >= 3 and sub[c1].std() > 0 and sub[c2].std() > 0:
                r, p = stats.pearsonr(sub[c1], sub[c2])
                corr_mat.loc[c1, c2] = r
                p_mat.loc[c1, c2] = p

    return corr_mat, p_mat, n_mat


def make_upper_triangle_table(corr_mat, p_mat=None, decimals=2,
                              sig_alpha=0.01, add_stars=False):
    """
    做成和论文表格类似的上三角形式。
    add_stars=True 时，控制台输出会用 ** 标记 99% 显著。
    """
    cols = corr_mat.columns.tolist()
    out = pd.DataFrame("", index=cols, columns=cols)

    for i, rname in enumerate(cols):
        for j, cname in enumerate(cols):
            if j < i:
                out.loc[rname, cname] = ""
                continue

            val = corr_mat.loc[rname, cname]

            if pd.isna(val):
                text = ""
            elif i == j:
                text = "1"
            else:
                if abs(val) < 0.5 * 10 ** (-decimals):
                    val = 0.0
                text = f"{val:.{decimals}f}"

                if add_stars and p_mat is not None:
                    p = p_mat.loc[rname, cname]
                    if pd.notna(p) and p <= sig_alpha:
                        text += "**"

            out.loc[rname, cname] = text

    return out


def export_corr_excel(corr_mat, p_mat, n_mat, out_path,
                      sheet_name="corr_upper_99sig",
                      title="表 1  EHCI 与潜在预报因子及其相互间的相关系数",
                      sig_alpha=0.01,
                      decimals=2,
                      meta_df=None):
    """
    导出 Excel：
    - corr_upper_99sig：论文式上三角表，99% 显著加粗
    - corr_full：完整相关系数矩阵
    - p_value：p 值矩阵
    - nobs：有效样本数矩阵
    - factor_meta：因子信息
    """
    upper = make_upper_triangle_table(
        corr_mat, p_mat=p_mat, decimals=decimals,
        sig_alpha=sig_alpha, add_stars=False
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        upper.to_excel(writer, sheet_name=sheet_name, startrow=2)
        corr_mat.to_excel(writer, sheet_name="corr_full")
        p_mat.to_excel(writer, sheet_name="p_value")
        n_mat.to_excel(writer, sheet_name="nobs")

        if meta_df is not None:
            meta_df.to_excel(writer, sheet_name="factor_meta", index=False)

        wb = writer.book
        ws = writer.sheets[sheet_name]

        nrow, ncol = upper.shape

        # 标题和说明
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol + 1)
        ws.cell(1, 1).value = title
        ws.cell(1, 1).font = Font(name="宋体", bold=True, size=13)
        ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol + 1)
        ws.cell(2, 1).value = "注：加粗代表通过 99% 显著性检验。"
        ws.cell(2, 1).font = Font(name="宋体", size=10)
        ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")

        # 样式
        thin = Side(style="thin", color="999999")
        border = Border(top=thin, bottom=thin, left=thin, right=thin)
        header_fill = PatternFill("solid", fgColor="EAEAEA")

        header_row = 3
        data_start_row = 4
        data_start_col = 2

        # 表头
        for cell in ws[header_row]:
            cell.font = Font(name="Times New Roman", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
            cell.border = border

        # 行名 + 数据区域
        for r in range(data_start_row, data_start_row + nrow):
            ws.cell(r, 1).font = Font(name="Times New Roman", bold=True, size=11)
            ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(r, 1).border = border

            for c in range(1, ncol + 2):
                ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(r, c).border = border

        # 99% 显著性加粗
        names = upper.index.tolist()
        for i, rname in enumerate(names):
            for j, cname in enumerate(names):
                excel_row = data_start_row + i
                excel_col = data_start_col + j

                if j < i or i == j:
                    continue

                p = p_mat.loc[rname, cname]
                if pd.notna(p) and p <= sig_alpha:
                    ws.cell(excel_row, excel_col).font = Font(
                        name="Times New Roman",
                        bold=True,
                        size=11
                    )
                else:
                    ws.cell(excel_row, excel_col).font = Font(
                        name="Times New Roman",
                        size=11
                    )

        # 冻结窗格
        ws.freeze_panes = "B4"

        # 列宽
        ws.column_dimensions["A"].width = 14
        for c in range(2, ncol + 2):
            ws.column_dimensions[get_column_letter(c)].width = 10

        # 行高
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 20

        # 其它 sheet 简单美化
        for sname in ["corr_full", "p_value", "nobs", "factor_meta"]:
            if sname in writer.sheets:
                w = writer.sheets[sname]
                w.freeze_panes = "B2"
                for col in range(1, min(w.max_column, 30) + 1):
                    w.column_dimensions[get_column_letter(col)].width = 12
                for row in w.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

    print(f"[Saved] {out_path}")


# ---------------------------------------------------------
# 9.1 整理相关分析数据
# ---------------------------------------------------------
# 按 MANUAL_FACTORS 的顺序输出因子
candidate_predictors = [f[0] for f in MANUAL_FACTORS if f[0] in X_train_all_dict]

# 训练期相关表：1962–2006
df_corr_train = pd.concat(
    [TS.rename("EHCI")] +
    [X_train_all_dict[x].rename(x) for x in candidate_predictors],
    axis=1
)

for c in df_corr_train.columns:
    df_corr_train[c] = pd.to_numeric(df_corr_train[c], errors="coerce")

df_corr_train = df_corr_train.replace([np.inf, -np.inf], np.nan)

corr_mat, p_mat, n_mat = calc_corr_p_n(df_corr_train)

# ---------------------------------------------------------
# 9.2 控制台打印：上三角相关系数表
# ---------------------------------------------------------
console_table = make_upper_triangle_table(
    corr_mat,
    p_mat=p_mat,
    decimals=2,
    sig_alpha=0.01,
    add_stars=True
)

print("\n表 1  EHCI 与潜在预报因子及其相互间的相关系数")
print("注：** 表示通过 99% 显著性检验。\n")
print(console_table.to_string())

# ---------------------------------------------------------
# 9.3 导出 Excel
# ---------------------------------------------------------
out_xlsx = fr"{PYFILE}/p5/data/manual_predictor_corr_table.xlsx"

export_corr_excel(
    corr_mat=corr_mat,
    p_mat=p_mat,
    n_mat=n_mat,
    out_path=out_xlsx,
    sheet_name="corr_upper_99sig",
    title="表 1  EHCI 与潜在预报因子及其相互间的相关系数",
    sig_alpha=0.01,
    decimals=2,
    meta_df=meta_df
)

print("Done. 不再画图。")

# 如果你没有删掉后面的绘图代码，可以保留这一句，防止继续往下执行
raise SystemExit

